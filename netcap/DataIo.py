# TODO 文件数据 数据 内存与磁盘的相互转换
from .config import RUNNING_CONFIG_KEYS,RUNNING_CONFIG_URLS,DATABASE_DIR,OUTPUT_DIR
from openpyxl import Workbook, load_workbook


class _DataIo:
    def __init__(self):
        self.output_dir = OUTPUT_DIR
        self.database_dir = DATABASE_DIR



class Txt:
    def __init__(self):
        self._urls_file = RUNNING_CONFIG_URLS.as_posix()
        self._keys_file = RUNNING_CONFIG_KEYS.as_posix()
        self._urls =None
        self._keys =None

    def get_urls(self):
        try:
            with open(self._urls_file,'r',encoding='utf-8') as f:
                self._urls = f.readlines()
            self._urls = [u.strip('\n') for u in self._urls]
            return self._urls
        except FileNotFoundError as e:
            raise e

    def get_keys(self):
        try:
            with open(self._keys_file,'r',encoding='utf-8') as f:
                self._keys = f.readlines()
            self._keys = [u.strip('\n') for u in self._keys]
            return self._keys
        except FileNotFoundError as e:
            raise e

class _Xlsx(_DataIo):
    def __init__(self):
        super().__init__()

class DyXlsx(_Xlsx):
    def __init__(self,excel_file:str=None):
        """
        抖音数据xlsx文件
        :param excel_file: xlsx文件路径
        :return:
        """
        super().__init__()
        self.xlsx_file = self.output_dir.joinpath('dy.xlsx') if excel_file is None else excel_file

        self._wb = Workbook()
        self.sheet = self._wb.active

    # def write_dict_list_to_excel(self, dict_list, field_order=None, append_mode=True):
    #     """
    #     将字典列表转换成二维列表并写入Excel
    #     :param dict_list: 字典列表
    #     :param field_order: 字段顺序列表，如果为None则使用字典的键作为字段顺序
    #     :param append_mode: 是否以追加模式写入数据，如果为True且文件存在，则从最后一行追加数据
    #     :return: None
    #     """
    #     if not dict_list:
    #         return
    #
    #     # 确定字段顺序
    #     if field_order is None:
    #         # 从第一个字典中获取所有键作为字段顺序
    #         field_order = list(dict_list[0].keys())
    #
    #     # 检查文件是否存在以及是否需要追加数据
    #     try:
    #         if append_mode and self.xlsx_file.exists():
    #             # 加载现有文件
    #             self._wb = load_workbook(self.xlsx_file)
    #             self.sheet = self._wb.active
    #             # 不需要再次写入表头，直接追加数据
    #         else:
    #             # 创建新的工作簿
    #             self._wb = Workbook()
    #             self.sheet = self._wb.active
    #             # 写入表头
    #             self.sheet.append(field_order)
    #     except Exception as e:
    #         # 发生异常时，创建新的工作簿
    #         print(f"加载现有文件时出错: {e}，将创建新文件")
    #         self._wb = Workbook()
    #         self.sheet = self._wb.active
    #         # 写入表头
    #         self.sheet.append(field_order)
    #
    #     # 遍历字典列表，将每个字典转换为行数据
    #     for item in dict_list:
    #         # 根据字段顺序提取每个字段的值
    #         row_data = [item.get(field, '') for field in field_order]
    #         self.sheet.append(row_data)
    #
    #     # 保存Excel文件
    #     self._wb.save(self.xlsx_file)

    def write_dict_list_to_excel(self, dict_list, field_order=None, append_mode=True):
        """
        将字典列表转换成二维列表并写入Excel
        :param dict_list: 字典列表
        :param field_order: 字段顺序列表，如果为None则使用字典的键作为字段顺序
        :param append_mode: 是否以追加模式写入数据，如果为True且文件存在，则从最后一行追加数据
        :return: None
        """
        if not dict_list:
            return

        # 确定字段顺序
        if field_order is None:
            # 从第一个字典中获取所有键作为字段顺序
            field_order = list(dict_list[0].keys())

        # 检查文件是否存在以及是否需要追加数据
        try:
            if append_mode and self.xlsx_file.exists():
                # 加载现有文件
                self._wb = load_workbook(self.xlsx_file)
                self.sheet = self._wb.active
                # 不需要再次写入表头，直接追加数据
            else:
                # 创建新的工作簿
                self._wb = Workbook()
                self.sheet = self._wb.active
                # 写入表头
                self.sheet.append(field_order)
        except Exception as e:
            # 发生异常时，创建新的工作簿
            print(f"加载现有文件时出错: {e}，将创建新文件")
            self._wb = Workbook()
            self.sheet = self._wb.active
            # 写入表头
            self.sheet.append(field_order)

        # 遍历字典列表，将每个字典转换为行数据
        for item in dict_list:
            # 根据字段顺序提取每个字段的值
            row_data = [item.get(field, '') for field in field_order]
            self.sheet.append(row_data)

        # 保存Excel文件
        self._wb.save(self.xlsx_file)

    def get(self,data):
        # TODO 此方法需要日后实现
        pass

    @staticmethod
    def test():
        print('test','*'*50)



class Db(_DataIo):
    def __init__(self):
        super().__init__()
        self._db_file = DATABASE_DIR.as_posix()


class DyDB(Db):
    def __init__(self):
        super().__init__()